import openpyxl as xl
import os
print(os.getcwd())
os.chdir('C:\DB')
print(os.getcwd())
exf = xl.load_workbook('ess.xlsx')
aws = exf.active
for i in aws.rows:
    index = i[0].row
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    mat = i[3].value

    tot = kor + eng + mat
    avg = tot / 3
    # column값은 1번 = A열
    aws.cell(row = index, column = 5).value = tot
    aws.cell(row = index, column = 6).value = round(avg, 2)

    print(f'{index} {name} {kor} {eng} {mat} {tot} {avg:.2f}')
exf.save('outss.xlsx')
exf.close()
