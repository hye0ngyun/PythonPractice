import openpyxl as xl

exf = xl.load_workbook('c://DB//itx.xlsx')
aws = exf.active
tot = 0
for i in aws.rows:
    tot += i[1].value
    print(f'i[0].value: {i[0].value}, i[1].value: {i[1].value}, len(aws.rows): {len(list(aws.rows))}')
avg = tot // len(list(aws.rows))
# print(list(aws.rows))
print(f'{avg}')
aws.cell(row = 6, column = 1).value = 'avg'
aws.cell(row = 6, column = 2).value = avg
exf.save('c://DB//outitx.xlsx')